"""
Read the Order details xlsx and extract data from it to the  Order classes
"""
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl import load_workbook, Workbook
from collections import OrderedDict

ORDER_START_TOKEN = 'Line Item:'
SIZE_BEGIN_TOKEN = 'Size'
SIZE_END_TOKEN = 'Total Qty:'


class StateError(Exception):
    def __init__(self, message):
        self.message = message

class get_order_detail():

    STATES = ["NO_ORDER", "STYLE_COLOR", "STYLE_NAME", "WAIT_SIZE", "SIZE", "FINISHED"]

    def __init__(self, CRDate=None, ord_id=None, wholesale=None, ord_name=None, sizes=None):
        self.state = "NO_ORDER"
        self.CRDate = CRDate
        self.ord_id = ord_id
        self.wholesale = wholesale
        self.ord_name = ord_name
        self.sizes = sizes

    def __call__(self, row):
        """
        gets a new row to parse
        """
        return None
        if self.state == "NO_ORDER":
            if row[0].value == ORDER_START_TOKEN:
                self.CRDate = row[4].value
                self.state = "STYLE_COLOR"
            return None

        if self.state == "STYLE_COLOR":
            self.ord_id = row[1].value
            self.wholesale = row[7].value
            self.state = "STYLE_NAME"
            return None

        if self.state == "STYLE_NAME":
            self.ord_name = row[1].value
            self.state = "WAIT_SIZE"
            return None

        if self.state == "WAIT_SIZE":
            if row[0].value == SIZE_BEGIN_TOKEN:
                self.sizes = OrderedDict([('S',0), ('M',0), ('L',0), ('XL',0)])
                self.state = "SIZE"
            return None

        if self.state == "SIZE":
            if row[0].value == SIZE_END_TOKEN:
                self.state = "FINISHED"
                return self
            self.sizes[row[0].value] = int(row[2].value)
            return None

        if self.state != "FINISHED":
            raise StateError("Undefined state found!: {state}".format(**self.__dict__))

    def __str__(self):
        return "{CRDate}, {ord_id}, {wholesale}, {ord_name}, {sizes}, {state}".format(**self.__dict__)

    def to_list(self):
        tmp = [self.ord_name, self.ord_id, self.wholesale, self.CRDate]
        tmp.extend([None, None, None])
        tmp.extend(self.sizes.values())
        return tmp



def get_sheet_width(ws):
    width = None
    cell_range = ws['A1':'AB1']
    for c in cell_range:
        for i in range(29):
            #print(c[i].value, end=', ')
            if str(c[i].value).startswith('=SUM'):
                return i
        #print()
    return width

def get_sheet_header(ws, ws_width):
    header = []
    cell_range = ws['A2':'AB2']
    for c in cell_range:
        for i in range(2,ws_width):
            #print(c[i].value, end=', ')
            header.append(c[i].value)
        #print()
    #print(header)
    return header

def get_xlsx_lines(ws):
    xlsx_lines = []
    j = 0
    ws_width = get_sheet_width(ws)
    header = get_sheet_header(ws, ws_width)
    #xlsx_lines.append(header)
    line = []
    for row in ws.rows:
        if row[2].fill and row[2].fill.start_color.index == "FFFF0000":
            pass
            #print(j, end=': ')
            #for i in range(2,ws_width):
            #    print(row[i].value, end=', ')
            #print()
        elif row[2].value == None or j == 1:
            pass
        else:
            for i in range(2,ws_width):
                line.append(row[i].value)
            xlsx_lines.append(line)
            line = []
        j += 1
        if j == 100:
            break
    return [header, xlsx_lines]
    
def read_xlsx( item_class, filename):
    wb = load_workbook(filename=filename, read_only=True)
    ws_names = wb.get_sheet_names()
    print(ws_names)
    order_list = []
    for ws_name in ws_names:
        if ws_name == 'Total Preorder':
            continue
        ws = wb[ws_name]
        tmp = get_xlsx_lines(ws)
        print(tmp)
        if tmp:
            order_list.append(tmp)
    return order_list

def sheet_to_entries(sheet, ws, row):
    header = sheet[0]
    j = 1
    for l in sheet[1]:
        i = 0
        for c in l[4:]:
            if c:
                ws['A{}'.format(j)] = str(l[0]) + "-" + str(header[i])
                #ws['A5'] = str(c[i])
                j += 1
            i += 1
    return row
            
        

def write_xlsx(orders):
    wb = Workbook()
    ws = wb.active
    row = 0
    for order in orders:
        row = sheet_to_entries(order, ws, row)
        #ws.append(order.to_list())
    
    wb.save('../test/out2.xlsx')

def main():
    order_list = read_xlsx(get_order_detail,
                           '../test/sg/Q1Adidas_Mother2.xlsx')
    write_xlsx(order_list)
    print("Conversion Finished")

#main()

if __name__ == "__main__":
    main()
