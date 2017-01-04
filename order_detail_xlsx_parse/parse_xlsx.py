# -*- coding: utf-8 -*-
"""
Read the Order details xlsx and extract data from it to the  Order classes
"""
from openpyxl import load_workbook, Workbook
from collections import OrderedDict

ORDER_START_TOKEN = 'Line Item:'
SIZE_BEGIN_TOKEN = 'Size'
SIZE_END_TOKEN = 'Total Qty:'


class StateError(Exception):
    def __init__(self, message):
        self.message = message

class get_order_detail():

    STATES = ["NO_ORDER", "STYLE_COLOR", "STYLE_NAME", "WAIT_SIZE", "SIZE", "FINISHED"]

    def __init__(self, CRDate=None, ord_id=None, wholesale=None, ord_name=None, sizes=None):
        self.state = "NO_ORDER"
        self.CRDate = CRDate
        self.ord_id = ord_id
        self.wholesale = wholesale
        self.ord_name = ord_name
        self.sizes = sizes

    def __call__(self, row):
        """
        gets a new row to parse
        """
        if self.state == "NO_ORDER":
            if row[0].value == ORDER_START_TOKEN:
                self.CRDate = row[4].value
                self.state = "STYLE_COLOR"
            return None

        if self.state == "STYLE_COLOR":
            self.ord_id = row[1].value
            self.wholesale = row[7].value
            self.state = "STYLE_NAME"
            return None

        if self.state == "STYLE_NAME":
            self.ord_name = row[1].value
            self.state = "WAIT_SIZE"
            return None

        if self.state == "WAIT_SIZE":
            if row[0].value == SIZE_BEGIN_TOKEN:
                self.sizes = OrderedDict([('S',0), ('M',0), ('L',0), ('XL',0)])
                self.state = "SIZE"
            return None

        if self.state == "SIZE":
            if row[0].value == SIZE_END_TOKEN:
                self.state = "FINISHED"
                return self
            self.sizes[row[0].value] = int(row[2].value)
            return None

        if self.state != "FINISHED":
            raise StateError("Undefined state found!: {state}".format(**self.__dict__))

    def __str__(self):
        return "{CRDate}, {ord_id}, {wholesale}, {ord_name}, {sizes}, {state}".format(**self.__dict__)

    def to_list(self):
        tmp = [self.ord_name, self.ord_id, self.wholesale, self.CRDate]
        tmp.extend([None, None, None])
        tmp.extend(self.sizes.values())
        return tmp

def write_xlsx(orders):
    wb = Workbook()
    ws = wb.active
    for order in orders:
        ws.append(order.to_list())
    wb.save('../test/out.xlsx')

def read_xlsx( item_class, filename):
    wb = load_workbook(filename=filename, read_only=True)
    ws_names = wb.get_sheet_names()
    order_list = []
    for ws_name in ws_names:
        ws = wb[ws_name]
        i = j = 0
        item_instance = item_class()
        for row in ws.rows:
            tmp = item_instance(row)
            if tmp:
                order_list.append(tmp)
                item_instance = item_class()
    return order_list

def main():
    order_list = read_xlsx(get_order_detail,
                           '../test/Nike_order_details_727856184_20170103.xlsx')
    write_xlsx(order_list)
    print("Conversion Finished")

main()

if __name__ == "__main__":
    main()
